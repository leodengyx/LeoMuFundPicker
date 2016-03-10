#!/usr/bin/python

import downloader
import logger
import mutualfund

import datetime
from openpyxl import Workbook
import threading

logger = logger.get_logger(__name__)


class XlsExporter:

    def __init__(self):

        self.xlsx_file_inst = Workbook()
        self.work_sheet_name = "Mutual Fund Data"
        self.next_avail_row = 1
        self.xlsx_file_lock = threading.Lock()

        for i in range(20):
            self.xlsx_file_inst.create_sheet(str(i))

    def write_mutual_fund_to_file(self, mutual_fund_inst, sheet_id, row_id):

        data_list = [mutual_fund_inst.fund_id,
                     mutual_fund_inst.fund_name,
                     mutual_fund_inst.fund_size,
                     mutual_fund_inst.mer,
                     mutual_fund_inst.status,
                     mutual_fund_inst.min_inve_initial,
                     mutual_fund_inst.category,
                     mutual_fund_inst.inve_style,
                     mutual_fund_inst.inve_objective_strategy,
                     mutual_fund_inst.growth_of_ten_thousand_YTD,
                     mutual_fund_inst.growth_of_ten_thousand_1month,
                     mutual_fund_inst.growth_of_ten_thousand_1year,
                     mutual_fund_inst.growth_of_ten_thousand_3year,
                     mutual_fund_inst.growth_of_ten_thousand_5year,
                     mutual_fund_inst.growth_of_ten_thousand_10year,
                     mutual_fund_inst.growth_fund_YTD,
                     mutual_fund_inst.growth_fund_1month,
                     mutual_fund_inst.growth_fund_1year,
                     mutual_fund_inst.growth_fund_3year,
                     mutual_fund_inst.growth_fund_5year,
                     mutual_fund_inst.growth_fund_10year,
                     mutual_fund_inst.growth_comp_index_YTD,
                     mutual_fund_inst.growth_comp_index_1month,
                     mutual_fund_inst.growth_comp_index_1year,
                     mutual_fund_inst.growth_comp_index_3year,
                     mutual_fund_inst.growth_comp_index_5year,
                     mutual_fund_inst.growth_comp_index_10year,
                     mutual_fund_inst.growth_comp_category_YTD,
                     mutual_fund_inst.growth_comp_category_1month,
                     mutual_fund_inst.growth_comp_category_1year,
                     mutual_fund_inst.growth_comp_category_3year,
                     mutual_fund_inst.growth_comp_category_5year,
                     mutual_fund_inst.growth_comp_category_10year]

        self.xlsx_file_lock.acquire()

        for i in range(len(data_list)):
            cell = self.xlsx_file_inst.get_sheet_by_name(str(sheet_id)).cell(row=row_id+1, column=i+1)
            cell.value = data_list[i]

        self.xlsx_file_lock.release()

        logger.debug("Write data: %s", data_list)

    def save_file(self, file_name=None):

        # Write Title
        work_sheet = self.xlsx_file_inst.create_sheet(self.work_sheet_name)
        title_list = ["FundId",
                      "FundName",
                      "FundSize(Mil)",
                      "MER(%)",
                      "Status",
                      "MinInvest",
                      "Category",
                      "InvestStyle",
                      "Object & Strategy",
                      "Growth of 1w YTD",
                      "Growth of 1w 1Month",
                      "Growth of 1w 1Year",
                      "Growth of 1w 3Year",
                      "Growth of 1w 5Year",
                      "Growth of 1w 10Year",
                      "Fund Growth YTD(%)",
                      "Fund Growth 1Month(%)",
                      "Fund Growth 1Year(%)",
                      "Fund Growth 3Year(%)",
                      "Fund Growth 5Year(%)",
                      "Fund Growth 10Year(%)",
                      "Index Growth YTD(%)",
                      "Index Growth 1Month(%)",
                      "Index Growth 1Year(%)",
                      "Index Growth 3Year(%)",
                      "Index Growth 5Year(%)",
                      "Index Growth 10Year(%)",
                      "Category Growth YTD(%)",
                      "Category Growth 1Month(%)",
                      "Category Growth 1Year(%)",
                      "Category Growth 3Year(%)",
                      "Category Growth 5Year(%)",
                      "Category Growth 10Year(%)"
                      ]
        for i in range(len(title_list)):
            cell = work_sheet.cell(row=self.next_avail_row, column=i+1)
            cell.value = title_list[i]
        logger.debug("Write title: %s", title_list)
        self.next_avail_row += 1

        # Combine data from all sheets
        for i in range(20):
            sheet = self.xlsx_file_inst.get_sheet_by_name(str(i))
            for row in sheet.rows:
                column = 1
                for cell in row:
                    work_sheet.cell(row=self.next_avail_row, column=column).value = cell.value
                    column += 1
                self.next_avail_row += 1

        # Remove old sheets
        for i in range(20):
            sheet = self.xlsx_file_inst.get_sheet_by_name(str(i))
            self.xlsx_file_inst.remove_sheet(sheet)

        if file_name == None:
            now = datetime.datetime.now()
            file_name = now.isoformat().replace(":","-") + ".xlsx"

        self.xlsx_file_inst.save(file_name);
        logger.debug("Save file: %s" % file_name)




